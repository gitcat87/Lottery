import sys
from PyQt5.QtWidgets import *
from PyQt5 import uic
from lotto4 import Workbook
from openpyxl import Workbook,load_workbook
import lotto3,lotto4, totalnum
import datetime


#주 메인 창(Form타입)
class FirstClass(QWidget) :
    def __init__(self) :
        super().__init__()
        self.ui = uic.loadUi("lotto/number.ui",self)

        self.num1.setText("")
        self.num2.setText("")
        self.num3.setText("")
        self.num4.setText("")
        self.num5.setText("")
        self.num6.setText("")
        self.bns1.setText("")

        DateTime= datetime.datetime.now()

        self.date.setText('오늘 날짜: %s년 %s월 %s일' %(DateTime.year, DateTime.month, DateTime.day))
        

        self.calllottery.clicked.connect(self.doA) #최근조회 실행
        self.numbutton.clicked.connect(self.doB) #직접입력 조회 실행
        self.data1.clicked.connect(self.doD) #통계조회 실행
        self.show() 

    #lotto4.py를 통해 저장한 lotto4.xlsx를 불러옵니다.
    def doA(self):
        lotto4.lot2(self)
        lw = load_workbook("lotto/lotto4.xlsx")
        ws = lw.active

        self.turn.setText(str(ws["B1"].value))
        self.num1.setText(str(ws["A2"].value))
        self.num2.setText(str(ws["B2"].value))
        self.num3.setText(str(ws["C2"].value))
        self.num4.setText(str(ws["D2"].value))
        self.num5.setText(str(ws["E2"].value))
        self.num6.setText(str(ws["F2"].value))
        self.bns1.setText(str(ws["H2"].value))
        self.potmoney.setText(str(ws["B3"].value))
    
    
    def doB(self):
        text = self.turnnum.text() #number.ui/ 회차 칸에 입력한 회차를 표기합니다
        print(text)

        wb = Workbook()
        ws = wb.active

        ws.append([text]) #직접 입력한 회차를 조회하기 위하여 엑셀 파일을 추가로 만듭니다
                          #lotto3.py의 def lot을 참조하세요

        wb.save("lotto/selfsearch.xlsx") #조회할 ""회차 숫자를"" 저장
        wb.close()

        lotto3.lot(self)
        FirstClass.doC(self) #엑셀 작성이 끝났으므로 다음 작업을 이어나갑니다

    def doC(self):
        lw = load_workbook("lotto/lotto4.xlsx") 
        ws = lw.active

        self.turn.setText(str(ws["B1"].value))
        self.num1.setText(str(ws["A2"].value))
        self.num2.setText(str(ws["B2"].value))
        self.num3.setText(str(ws["C2"].value))
        self.num4.setText(str(ws["D2"].value))
        self.num5.setText(str(ws["E2"].value))
        self.num6.setText(str(ws["F2"].value))
        self.bns1.setText(str(ws["H2"].value))
        self.potmoney.setText(str(ws["B3"].value)) 
    
    # 통계 창을 새로 띄우기 위함입니다
    def doD(self):
        totalnum.lot3(self) #통계 데이타 작성 함수 실행
        new = Second()     #두 번째 창 출력
        new.show()


       
# 통계 창(Form 타입)
class Second(QWidget):
    def __init__(self) :
        super().__init__()
        self.ui = uic.loadUi("lotto/legend.ui",self)
        lw = load_workbook("lotto/totalnum.xlsx")
        ws = lw.active

        self.legend_num1.setText(str(ws["A2"].value))
        self.legend_num2.setText(str(ws["A3"].value))

        self.appear_num1.setText(str(ws["A4"].value))
        self.appear_num2.setText(str(ws["B4"].value))
        self.appear_num3.setText(str(ws["C4"].value))
        self.appear_num4.setText(str(ws["D4"].value))
        self.appear_num5.setText(str(ws["E4"].value))


        self.show()
        


if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = FirstClass() 
    myWindow.show()
    app.exec_()
